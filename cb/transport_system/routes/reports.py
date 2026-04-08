import io
import os
from flask import Blueprint, render_template, request, jsonify, send_file, make_response
from models import db, Academy, Team, Vehicle, Trip
from algorithm import get_distribution

reports_bp = Blueprint('reports', __name__)

# تسجيل الخط العربي مع reportlab عند بدء التشغيل
_font_registered = False

def _register_arabic_fonts():
    global _font_registered
    if _font_registered:
        return
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        basedir = os.path.abspath(os.path.dirname(__file__))
        basedir = os.path.abspath(os.path.join(basedir, '..'))
        font_dir = os.path.join(basedir, 'static', 'fonts')
        pdfmetrics.registerFont(TTFont('Tahoma', os.path.join(font_dir, 'tahoma.ttf')))
        pdfmetrics.registerFont(TTFont('TahomaBold', os.path.join(font_dir, 'tahomabd.ttf')))
        _font_registered = True
    except Exception as e:
        print(f"Warning: Could not register Arabic fonts: {e}")


@reports_bp.route('/reports')
def reports_page():
    """صفحة التقارير والتصدير"""
    academies = Academy.query.order_by(Academy.name).all()
    return render_template('reports.html', academies=academies)


@reports_bp.route('/api/reports/preview', methods=['POST'])
def api_preview_report():
    """API: معاينة التقرير"""
    data = request.get_json()
    report_type = data.get('type', 'general')
    academy_id = data.get('academy_id')
    day = data.get('day', 0)

    if report_type == 'academy' and not academy_id:
        return jsonify({'error': 'يجب اختيار الأكاديمية'}), 400

    result = _build_report_data(report_type, academy_id, day)
    return jsonify(result)


@reports_bp.route('/reports/pdf')
def export_pdf():
    """تصدير تقرير PDF"""
    report_type = request.args.get('type', 'general')
    academy_id = request.args.get('academy_id', type=int)
    day = request.args.get('day', 0, type=int)

    report_data = _build_report_data(report_type, academy_id, day)

    # تسجيل الخط العربي
    _register_arabic_fonts()

    html = render_template('report_pdf.html', report=report_data,
                           font_regular='', font_bold='')

    try:
        from xhtml2pdf import pisa

        output = io.BytesIO()
        pisa_status = pisa.CreatePDF(html, dest=output)

        if pisa_status.err:
            return jsonify({'error': 'حدث خطأ أثناء توليد PDF'}), 500

        output.seek(0)
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=transport_report.pdf'
        return response
    except ImportError:
        return jsonify({'error': 'مكتبة xhtml2pdf غير مثبتة. يرجى تثبيتها أولاً.'}), 500
    except Exception as e:
        return jsonify({'error': f'خطأ في توليد PDF: {str(e)}'}), 500


@reports_bp.route('/reports/excel')
def export_excel():
    """تصدير تقرير Excel"""
    report_type = request.args.get('type', 'general')
    academy_id = request.args.get('academy_id', type=int)
    day = request.args.get('day', 0, type=int)

    report_data = _build_report_data(report_type, academy_id, day)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير النقل"
        ws.sheet_view.rightToLeft = True

        # أنماط
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='C2185B', end_color='C2185B', fill_type='solid')
        sub_header_font = Font(name='Arial', size=11, bold=True)
        sub_header_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
        cell_font = Font(name='Arial', size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # العنوان
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        title_cell = ws.cell(row=row, column=1)
        title_cell.value = report_data.get('title', 'تقرير النقل')
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = center_align
        row += 2

        # لكل عربة
        for v_data in report_data.get('vehicles_data', []):
            vehicle = v_data.get('vehicle', {})
            trips = v_data.get('trips', [])

            # اسم العربة
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            v_cell = ws.cell(row=row, column=1)
            v_cell.value = f"العربة: {vehicle.get('name', '')}"
            v_cell.font = sub_header_font
            v_cell.fill = sub_header_fill
            v_cell.alignment = center_align
            row += 1

            # رؤوس الأعمدة
            headers = ['الرحلة', 'الاتجاه', 'الأكاديمية', 'الفريق', 'الفئة']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = h
                cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='E91E63', end_color='E91E63', fill_type='solid')
                cell.alignment = center_align
                cell.border = border
            row += 1

            for trip in trips:
                for team in trip.get('teams', []):
                    cells_data = [
                        trip.get('trip_order', ''),
                        trip.get('direction_label', ''),
                        team.get('academy_name', ''),
                        team.get('gender', ''),
                        team.get('category_label', '')
                    ]
                    for col, val in enumerate(cells_data, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = str(val)
                        cell.font = cell_font
                        cell.alignment = center_align
                        cell.border = border
                    row += 1

            row += 1

        # ضبط عرض الأعمدة
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='transport_report.xlsx'
        )

    except ImportError:
        return jsonify({'error': 'مكتبة openpyxl غير مثبتة'}), 500
    except Exception as e:
        return jsonify({'error': f'خطأ في توليد Excel: {str(e)}'}), 500


def _build_report_data(report_type, academy_id, day):
    """بناء بيانات التقرير"""
    days = [1, 2] if day == 0 else [day]
    vehicles_data = []
    title = 'البطولة الوطنية لكرة السلة بكلميم — تقرير النقل'
    academy_sections = []

    if report_type == 'all_academies':
        title = 'البطولة الوطنية لكرة السلة بكلميم — تقرير جميع الأكاديميات'
        all_academies = Academy.query.order_by(Academy.name).all()

        for academy in all_academies:
            acad_vehicles = []
            for d in days:
                dist = get_distribution(d)
                for v_id, v_data in dist.get('vehicles', {}).items():
                    relevant_trips = []
                    for trip in v_data.get('trips', []):
                        relevant_teams = [t for t in trip.get('teams', []) if t.get('academy_id') == academy.id]
                        if relevant_teams:
                            trip_copy = dict(trip)
                            trip_copy['teams'] = relevant_teams
                            relevant_trips.append(trip_copy)
                    if relevant_trips:
                        acad_vehicles.append({
                            'vehicle': v_data['vehicle'],
                            'trips': relevant_trips,
                            'day': d
                        })

            if acad_vehicles:
                academy_sections.append({
                    'academy_name': academy.name,
                    'academy_hotel': academy.hotel,
                    'vehicles_data': acad_vehicles
                })
                vehicles_data.extend(acad_vehicles)

    elif report_type == 'academy' and academy_id:
        academy = Academy.query.get(academy_id)
        if not academy:
            return {'error': 'الأكاديمية غير موجودة'}
        title = f'البطولة الوطنية لكرة السلة بكلميم — {academy.name}'

        for d in days:
            dist = get_distribution(d)
            for v_id, v_data in dist.get('vehicles', {}).items():
                relevant_trips = []
                for trip in v_data.get('trips', []):
                    relevant_teams = [t for t in trip.get('teams', []) if t.get('academy_id') == academy_id]
                    if relevant_teams:
                        trip_copy = dict(trip)
                        trip_copy['teams'] = relevant_teams
                        relevant_trips.append(trip_copy)
                if relevant_trips:
                    vehicles_data.append({
                        'vehicle': v_data['vehicle'],
                        'trips': relevant_trips,
                        'day': d
                    })
    else:
        for d in days:
            dist = get_distribution(d)
            for v_id, v_data in dist.get('vehicles', {}).items():
                vehicles_data.append({
                    'vehicle': v_data['vehicle'],
                    'trips': v_data.get('trips', []),
                    'day': d
                })

    day_label = 'كلا اليومين' if day == 0 else f'اليوم {day}'

    return {
        'title': title,
        'day_label': day_label,
        'report_type': report_type,
        'academy_id': academy_id,
        'vehicles_data': vehicles_data,
        'total_vehicles': len(vehicles_data),
        'academy_sections': academy_sections
    }
