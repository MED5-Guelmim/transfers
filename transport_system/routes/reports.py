import io
import os
from flask import Blueprint, render_template, request, jsonify, send_file, make_response
from flask_login import login_required
from models import db, Academy, Team, Vehicle, Trip
from algorithm import get_distribution

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/reports')
@login_required
def reports_page():
    """صفحة التقارير والتصدير"""
    academies = Academy.query.order_by(Academy.name).all()
    return render_template('reports.html', academies=academies)


@reports_bp.route('/api/reports/preview', methods=['POST'])
@login_required
def api_preview_report():
    """API: معاينة التقرير"""
    data = request.get_json()
    report_type = data.get('type', 'general')  # general أو academy
    academy_id = data.get('academy_id')
    day = data.get('day', 0)  # 0 = كلاهما

    if report_type == 'academy' and not academy_id:
        return jsonify({'error': 'يجب اختيار الأكاديمية'}), 400

    result = _build_report_data(report_type, academy_id, day)
    return jsonify(result)


# ═══════════════════════════════════════════════════════════
# أدوات العربية لـ ReportLab
# ═══════════════════════════════════════════════════════════

def _reshape_arabic(text):
    """تشكيل وترتيب النص العربي لعرضه بشكل صحيح في PDF"""
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    except Exception:
        return str(text)


def _get_font_path():
    """الحصول على مسار ملفات الخطوط"""
    basedir = os.path.abspath(os.path.dirname(__file__))
    basedir = os.path.abspath(os.path.join(basedir, '..'))
    font_dir = os.path.join(basedir, 'static', 'fonts')
    return {
        'regular': os.path.join(font_dir, 'tahoma.ttf'),
        'bold': os.path.join(font_dir, 'tahomabd.ttf')
    }


def _register_arabic_fonts():
    """تسجيل الخطوط العربية في ReportLab"""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    fonts = _get_font_path()
    try:
        pdfmetrics.registerFont(TTFont('Arabic', fonts['regular']))
        pdfmetrics.registerFont(TTFont('ArabicBold', fonts['bold']))
    except Exception:
        # إذا لم تُوجد الخطوط، استخدم Helvetica
        pass


# ═══════════════════════════════════════════════════════════
# تصدير PDF بـ ReportLab
# ═══════════════════════════════════════════════════════════

@reports_bp.route('/reports/pdf')
@login_required
def export_pdf():
    """تصدير تقرير PDF باستخدام ReportLab مع دعم العربية الكامل"""
    report_type = request.args.get('type', 'general')
    academy_id = request.args.get('academy_id', type=int)
    day = request.args.get('day', 0, type=int)

    report_data = _build_report_data(report_type, academy_id, day)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        )
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT

        _register_arabic_fonts()

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=15 * mm,
            leftMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm
        )

        # ─── أنماط النصوص ───
        font_name = 'Arabic'
        font_bold = 'ArabicBold'

        # التأكد من وجود الخطوط
        from reportlab.pdfbase import pdfmetrics
        available_fonts = pdfmetrics.getRegisteredFontNames()
        if 'Arabic' not in available_fonts:
            font_name = 'Helvetica'
            font_bold = 'Helvetica-Bold'

        title_style = ParagraphStyle(
            'TitleStyle',
            fontName=font_bold,
            fontSize=16,
            alignment=TA_CENTER,
            textColor=colors.white,
            spaceAfter=2 * mm,
        )

        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            fontName=font_name,
            fontSize=11,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#fce4ec'),
            spaceAfter=5 * mm,
        )

        vehicle_header_style = ParagraphStyle(
            'VehicleHeader',
            fontName=font_bold,
            fontSize=12,
            alignment=TA_RIGHT,
            textColor=colors.HexColor('#880e4f'),
            spaceAfter=2 * mm,
        )

        cell_style = ParagraphStyle(
            'CellStyle',
            fontName=font_name,
            fontSize=9,
            alignment=TA_RIGHT,
            leading=13,
        )

        cell_center_style = ParagraphStyle(
            'CellCenterStyle',
            fontName=font_name,
            fontSize=9,
            alignment=TA_CENTER,
            leading=13,
        )

        header_cell_style = ParagraphStyle(
            'HeaderCellStyle',
            fontName=font_bold,
            fontSize=9,
            alignment=TA_CENTER,
            textColor=colors.white,
            leading=13,
        )

        academy_section_style = ParagraphStyle(
            'AcademySectionStyle',
            fontName=font_bold,
            fontSize=13,
            alignment=TA_RIGHT,
            textColor=colors.white,
            spaceAfter=3 * mm,
        )

        # ─── بناء المحتوى ───
        elements = []

        # ═══ العنوان ═══
        title_table_data = [
            [Paragraph(_reshape_arabic(report_data.get('title', 'تقرير النقل')), title_style)],
            [Paragraph(_reshape_arabic(report_data.get('day_label', '')), subtitle_style)]
        ]
        title_table = Table(title_table_data, colWidths=[doc.width])
        title_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#e91e63')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
        ]))
        elements.append(title_table)
        elements.append(Spacer(1, 5 * mm))

        # ═══ محتوى التقرير ═══
        if report_data.get('academy_sections'):
            # تقرير جميع الأكاديميات
            for section in report_data['academy_sections']:
                # عنوان الأكاديمية
                acad_header_data = [[Paragraph(
                    _reshape_arabic(f"{section['academy_name']} — {section['academy_hotel']}"),
                    academy_section_style
                )]]
                acad_header_table = Table(acad_header_data, colWidths=[doc.width])
                acad_header_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#880e4f')),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ]))
                elements.append(acad_header_table)
                elements.append(Spacer(1, 3 * mm))

                for v_data in section.get('vehicles_data', []):
                    _add_vehicle_table_to_elements(
                        elements, v_data, doc,
                        vehicle_header_style, header_cell_style,
                        cell_style, cell_center_style,
                        font_bold=font_bold,
                        is_academy_report=True
                    )

                elements.append(HRFlowable(
                    width="100%", thickness=1,
                    color=colors.HexColor('#f48fb1'),
                    spaceAfter=5 * mm, spaceBefore=5 * mm
                ))
        else:
            # تقرير عام أو أكاديمية واحدة
            is_academy = report_data.get('report_type') == 'academy'
            for v_data in report_data.get('vehicles_data', []):
                _add_vehicle_table_to_elements(
                    elements, v_data, doc,
                    vehicle_header_style, header_cell_style,
                    cell_style, cell_center_style,
                    font_bold=font_bold,
                    is_academy_report=is_academy
                )

        if not report_data.get('vehicles_data'):
            no_data_style = ParagraphStyle(
                'NoData', fontName=font_name, fontSize=12,
                alignment=TA_CENTER, textColor=colors.gray
            )
            elements.append(Spacer(1, 20 * mm))
            elements.append(Paragraph(_reshape_arabic('لا توجد بيانات للتقرير المحدد'), no_data_style))

        # ═══ التذييل ═══
        elements.append(Spacer(1, 10 * mm))
        footer_table_data = [[Paragraph(
            _reshape_arabic('البطولة الوطنية لكرة السلة بكلميم — نظام إدارة النقل اللوجستي'),
            ParagraphStyle('Footer', fontName=font_name, fontSize=8,
                           alignment=TA_CENTER, textColor=colors.gray)
        )]]
        footer_table = Table(footer_table_data, colWidths=[doc.width])
        footer_table.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (-1, 0), 2, colors.HexColor('#e91e63')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(footer_table)

        doc.build(elements)
        output.seek(0)

        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=transport_report.pdf'
        return response

    except ImportError as e:
        return jsonify({'error': f'مكتبة مطلوبة غير مثبتة: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'خطأ في توليد PDF: {str(e)}'}), 500


def _add_vehicle_table_to_elements(elements, v_data, doc,
                                    vehicle_header_style, header_cell_style,
                                    cell_style, cell_center_style,
                                    font_bold='ArabicBold',
                                    is_academy_report=False):
    """إضافة جدول عربة واحدة إلى عناصر PDF — الفرق مجمّعة في سطر واحد لكل رحلة"""
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer

    vehicle = v_data.get('vehicle', {})
    trips = v_data.get('trips', [])
    day_label = f" — اليوم {v_data.get('day', '')}" if v_data.get('day') else ''

    # اسم العربة
    vehicle_text = _reshape_arabic(
        f"العربة: {vehicle.get('name', '')} ({vehicle.get('type_label', '')}){day_label}"
    )
    v_header_data = [[Paragraph(vehicle_text, vehicle_header_style)]]
    v_header_table = Table(v_header_data, colWidths=[doc.width])
    v_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fce4ec')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(v_header_table)

    # ─── الجدول ───
    if is_academy_report:
        headers = ['الرحلة', 'الاتجاه', 'التوقيت', 'الفرق المنقولة']
        col_widths = [doc.width * 0.12, doc.width * 0.12, doc.width * 0.15, doc.width * 0.61]
    else:
        headers = ['الرحلة', 'الاتجاه', 'التوقيت', 'الفرق المنقولة']
        col_widths = [doc.width * 0.10, doc.width * 0.10, doc.width * 0.13, doc.width * 0.67]

    header_row = [Paragraph(_reshape_arabic(h), header_cell_style) for h in headers]
    table_data = [header_row]

    for trip in trips:
        # ★ تجميع كل الفرق في خلية واحدة (مثل المعاينة)
        if is_academy_report:
            teams_str = '، '.join([
                f"{t.get('gender', '')} {t.get('category_label', '')}"
                for t in trip.get('teams', [])
            ])
        else:
            teams_str = '، '.join([
                f"{t.get('academy_name', '')}: {t.get('gender', '')} {t.get('category_label', '')}"
                for t in trip.get('teams', [])
            ])

        time_display = trip.get('time_display', '')
        match_name = trip.get('match_name', '')
        trip_label = time_display if time_display else str(trip.get('trip_order', ''))

        direction_label = trip.get('direction_label', '')

        row = [
            Paragraph(_reshape_arabic(trip_label), cell_center_style),
            Paragraph(_reshape_arabic(direction_label), cell_center_style),
            Paragraph(_reshape_arabic(match_name), cell_center_style),
            Paragraph(_reshape_arabic(teams_str or '—'), cell_style),
        ]
        table_data.append(row)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    style_commands = [
        # الرأس
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c2185b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # الخطوط
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#f8bbd0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]

    # تلوين الصفوف بالتناوب
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fce4ec')))

    table.setStyle(TableStyle(style_commands))
    elements.append(table)
    elements.append(Spacer(1, 5 * mm))


# ═══════════════════════════════════════════════════════════
# تصدير Excel — نفس تنسيق المعاينة (فرق مجمّعة لكل رحلة)
# ═══════════════════════════════════════════════════════════

@reports_bp.route('/reports/excel')
@login_required
def export_excel():
    """تصدير تقرير Excel — الفرق مجمّعة في خلية واحدة لكل رحلة"""
    report_type = request.args.get('type', 'general')
    academy_id = request.args.get('academy_id', type=int)
    day = request.args.get('day', 0, type=int)

    report_data = _build_report_data(report_type, academy_id, day)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير النقل"
        ws.sheet_view.rightToLeft = True

        # أنماط
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='C2185B', end_color='C2185B', fill_type='solid')
        sub_header_font = Font(name='Arial', size=11, bold=True)
        sub_header_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
        cell_font = Font(name='Arial', size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ألوان الاتجاه
        go_font = Font(name='Arial', size=10, bold=True, color='2E7D32')
        return_font = Font(name='Arial', size=10, bold=True, color='1565C0')

        row = 1

        # العنوان
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        title_cell = ws.cell(row=row, column=1)
        title_cell.value = report_data.get('title', 'تقرير النقل')
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = center_align
        row += 2

        is_academy = report_data.get('report_type') in ['academy', 'all_academies']

        # ═══ لتقرير جميع الأكاديميات ═══
        if report_data.get('academy_sections'):
            for section in report_data['academy_sections']:
                # عنوان الأكاديمية
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                acad_cell = ws.cell(row=row, column=1)
                acad_cell.value = f"{section['academy_name']} — {section['academy_hotel']}"
                acad_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                acad_cell.fill = PatternFill(start_color='880E4F', end_color='880E4F', fill_type='solid')
                acad_cell.alignment = center_align
                row += 1

                for v_data in section.get('vehicles_data', []):
                    row = _write_vehicle_to_excel(
                        ws, v_data, row, is_academy_report=True,
                        sub_header_font=sub_header_font, sub_header_fill=sub_header_fill,
                        cell_font=cell_font, center_align=center_align,
                        right_align=right_align, border=border,
                        go_font=go_font, return_font=return_font
                    )
                row += 1
        else:
            # تقرير عام أو أكاديمية واحدة
            for v_data in report_data.get('vehicles_data', []):
                row = _write_vehicle_to_excel(
                    ws, v_data, row, is_academy_report=is_academy,
                    sub_header_font=sub_header_font, sub_header_fill=sub_header_fill,
                    cell_font=cell_font, center_align=center_align,
                    right_align=right_align, border=border,
                    go_font=go_font, return_font=return_font
                )

        # ضبط عرض الأعمدة
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 50

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='transport_report.xlsx'
        )

    except ImportError:
        return jsonify({'error': 'مكتبة openpyxl غير مثبتة'}), 500
    except Exception as e:
        return jsonify({'error': f'خطأ في توليد Excel: {str(e)}'}), 500


def _write_vehicle_to_excel(ws, v_data, row, is_academy_report,
                             sub_header_font, sub_header_fill,
                             cell_font, center_align, right_align, border,
                             go_font, return_font):
    """كتابة بيانات عربة واحدة في Excel — الفرق مجمّعة لكل رحلة"""
    from openpyxl.styles import Font, PatternFill

    vehicle = v_data.get('vehicle', {})
    trips = v_data.get('trips', [])
    day_label = f" — اليوم {v_data.get('day', '')}" if v_data.get('day') else ''

    # اسم العربة
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    v_cell = ws.cell(row=row, column=1)
    v_cell.value = f"العربة: {vehicle.get('name', '')} ({vehicle.get('type_label', '')}){day_label}"
    v_cell.font = sub_header_font
    v_cell.fill = sub_header_fill
    v_cell.alignment = center_align
    row += 1

    # رؤوس الأعمدة
    headers = ['الرحلة', 'الاتجاه', 'التوقيت', 'المباراة', 'الفرق المنقولة']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = h
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='E91E63', end_color='E91E63', fill_type='solid')
        cell.alignment = center_align
        cell.border = border
    row += 1

    # ★ كل رحلة في سطر واحد — الفرق مجمّعة
    for trip in trips:
        if is_academy_report:
            teams_str = '، '.join([
                f"{t.get('gender', '')} {t.get('category_label', '')}"
                for t in trip.get('teams', [])
            ])
        else:
            teams_str = '، '.join([
                f"{t.get('academy_name', '')}: {t.get('gender', '')} {t.get('category_label', '')}"
                for t in trip.get('teams', [])
            ])

        time_display = trip.get('time_display', '')
        match_name = trip.get('match_name', '')
        trip_label = time_display if time_display else str(trip.get('trip_order', ''))

        cells_data = [trip_label, trip.get('direction_label', ''), match_name, '', teams_str]

        # ← عمود المباراة هو الرابع والفرق هي الخامسة
        cells_data = [trip_label, trip.get('direction_label', ''), time_display, match_name, teams_str]

        for col, val in enumerate(cells_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = str(val)
            cell.font = cell_font
            cell.alignment = center_align if col <= 4 else right_align
            cell.border = border

            # تلوين الاتجاه
            if col == 2:
                if trip.get('direction') == 'go':
                    cell.font = go_font
                else:
                    cell.font = return_font

        row += 1

    row += 1  # فاصل
    return row


# ═══════════════════════════════════════════════════════════
# بناء بيانات التقرير
# ═══════════════════════════════════════════════════════════

def _build_report_data(report_type, academy_id, day):
    """بناء بيانات التقرير"""
    days = [1, 2] if day == 0 else [day]
    vehicles_data = []
    title = 'البطولة الوطنية لكرة السلة بكلميم — تقرير النقل'
    academy_sections = []  # لتقرير جميع الأكاديميات

    if report_type == 'all_academies':
        title = 'البطولة الوطنية لكرة السلة بكلميم — تقرير جميع الأكاديميات'
        all_academies = Academy.query.order_by(Academy.name).all()

        for academy in all_academies:
            acad_vehicles = []
            for d in days:
                dist = get_distribution(d)
                for v_id, v_data in dist.get('vehicles', {}).items():
                    relevant_trips = []
                    for trip in v_data.get('trips', []):
                        relevant_teams = [t for t in trip.get('teams', []) if t.get('academy_id') == academy.id]
                        if relevant_teams:
                            trip_copy = dict(trip)
                            trip_copy['teams'] = relevant_teams
                            relevant_trips.append(trip_copy)
                    if relevant_trips:
                        acad_vehicles.append({
                            'vehicle': v_data['vehicle'],
                            'trips': relevant_trips,
                            'day': d
                        })

            if acad_vehicles:
                academy_sections.append({
                    'academy_name': academy.name,
                    'academy_hotel': academy.hotel_name,
                    'vehicles_data': acad_vehicles
                })
                vehicles_data.extend(acad_vehicles)

    elif report_type == 'academy' and academy_id:
        academy = Academy.query.get(academy_id)
        if not academy:
            return {'error': 'الأكاديمية غير موجودة'}
        title = f'البطولة الوطنية لكرة السلة بكلميم — {academy.name}'

        for d in days:
            dist = get_distribution(d)
            for v_id, v_data in dist.get('vehicles', {}).items():
                relevant_trips = []
                for trip in v_data.get('trips', []):
                    relevant_teams = [t for t in trip.get('teams', []) if t.get('academy_id') == academy_id]
                    if relevant_teams:
                        trip_copy = dict(trip)
                        trip_copy['teams'] = relevant_teams
                        relevant_trips.append(trip_copy)
                if relevant_trips:
                    vehicles_data.append({
                        'vehicle': v_data['vehicle'],
                        'trips': relevant_trips,
                        'day': d
                    })
    else:
        for d in days:
            dist = get_distribution(d)
            for v_id, v_data in dist.get('vehicles', {}).items():
                vehicles_data.append({
                    'vehicle': v_data['vehicle'],
                    'trips': v_data.get('trips', []),
                    'day': d
                })

    day_label = 'كلا اليومين' if day == 0 else f'اليوم {day}'

    return {
        'title': title,
        'day_label': day_label,
        'report_type': report_type,
        'academy_id': academy_id,
        'vehicles_data': vehicles_data,
        'total_vehicles': len(vehicles_data),
        'academy_sections': academy_sections
    }
